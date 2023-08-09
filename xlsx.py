import openpyxl
from selenium import webdriver
from datetime import date
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
# os.chmod('/Users/sundar.g/Downloads/chromedriver_mac64/chromedriver', 755)
options = Options()
ser = Service("/Users/sundar.g/Downloads/chromedriver_mac64/chromedriver")
op = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=ser, options=op)
today = date.today()
year_=today.year
years=[]
empt_dict={}
for i in range(0,5):
    years.append(year_-i)
def citi_func_calling(url1,url2,url3):
    list_=[]
    value1=scopus_citi(url1)
    list_.append(value1)
    list_.append(web_of_science_citi(url2))
    list_.append(Google_scholar_citi(url3))
  

    return list_
def excel_read_write_fun(path):
    wb_obj = openpyxl.load_workbook(path)
    my_sheet_obj = wb_obj.active
    sheet = wb_obj["Sheet1"]
    print(my_sheet_obj.max_column)
    print(len(sheet["A"]))
    final_data={}
    for i in range(2,35):
        list_=[]
        name_cell_obj = my_sheet_obj.cell(row = i, column = 2).value
        print(name_cell_obj)
        scopus_cell_obj = my_sheet_obj.cell(row = i, column = 3).value
        
        Wos_cell_obj = my_sheet_obj.cell(row = i, column = 4).value
        googleScholar_cell_obj = my_sheet_obj.cell(row = i, column = 6).value
        list_.append(scopus_citi(scopus_cell_obj))
        # list_.append(web_of_science_citi(Wos_cell_obj))
        list_.append(Google_scholar_citi(googleScholar_cell_obj))
        final_data[name_cell_obj]=list_
        # print(final_data)
    return final_data
def Google_scholar_citi(url):
    if(url==None):
        for i in range(len(years)):
            empt_dict[str(years[i])]="0"
        return {"gs":empt_dict}
    else:
        url=f"https://scholar.google.co.in/citations?user={url}"
        driver.get(url)
        years_li=[]
        year_li2=[]
        dict1={}
        years_ele=driver.find_elements(By.XPATH,"""//*[local-name()='span' and @class="gsc_g_t"]""")
        for i in years_ele:
            year_li2.append(i.get_attribute("innerHTML"))
        # print(year_li2)
        years_val_ele=driver.find_elements(By.XPATH,"""//*[local-name()='a' and @class="gsc_g_a"]""")

        for i in range(1,len(years_val_ele)):

            value=driver.find_element(By.XPATH,f"""//*[local-name()='a' and @class="gsc_g_a"]{[i]}""").get_attribute("style").split(";")
            # print(value)
            value_i=value[3].split(":")
            # print(value_i)
            # print(int(value_i[1]))
            value2=driver.find_element(By.XPATH,f"""//*[local-name()='a' and @class="gsc_g_a"]{[i+1]}""").get_attribute("style").split(";")
            value_i2=value2[3].split(":")
            # print(int(value_i2[1]))
            
            if((int(value_i[1])-int(value_i2[1]))==1):
                    years_val_=driver.find_element(By.XPATH,f"""//*[@id="gsc_rsb_cit"]/div/div[3]/div/a{[i]}/span""").get_attribute("innerHTML") 
                    years_li.append(years_val_)
            else:
                years_val_=driver.find_element(By.XPATH,f"""//*[@id="gsc_rsb_cit"]/div/div[3]/div/a{[i]}/span""").get_attribute("innerHTML") 
                years_li.append(years_val_)
                n=int(value_i[1])-int(value_i2[1])
                for j in range(n-1):

                    years_li.append(0)

        if(len(years_val_ele)>0):

            years_val_=driver.find_element(By.XPATH,f"""//*[@id="gsc_rsb_cit"]/div/div[3]/div/a{[i+1]}/span""").get_attribute("innerHTML")
            years_li.append(years_val_)
        
        for i in range(len(years_li)):
            dict1[year_li2[i]]=years_li[i]
        for i in years:
            if str(i) not in (dict1.keys()):
                dict1[str(i)]=0
        keys_=list(dict1.keys())
        for i in keys_:
            if(int(i) not in years):
                del dict1[i]
        return {"gs":dict1}

def scopus_citi(url):
    print(url)
    if(url==None):
        for i in range(len(years)):
            empt_dict[str(years[i])]="0"
        return {"scopus":empt_dict}
    else:
        url1=int(url)
        url=f"https://www.scopus.com/authid/detail.uri?authorId={url1}"
        driver.get(url)
        try:
            e=WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH,"""//*[local-name()='path' and @class="highcharts-point"]""")))
            click_on_python=driver.find_elements(By.XPATH,f"""//*[local-name()='path'and @class="highcharts-halo highcharts-color-undefined" or local-name()='path'and@class="highcharts-point"]""")
            # print(click_on_python)
            length1=0
            dict_={}
            for i in click_on_python:
                # print(i.get_attribute("aria-label"))
                list_=i.get_attribute("aria-label").split(" ")
                dict_[list_[0][:4]]=list_[1]
                length1+=1

            print(length1)


            # dict_={}
            # for i in range(length1-4,length1+1):
            #     click_on_python=driver.find_element(By.XPATH,f"""//*[local-name()='path' and @class="highcharts-point"]{[i]}""")
            #     list_=click_on_python.get_attribute("aria-label").split(" ")
            #     dict_[list_[0][:4]]=list_[1]
            # print(dict_)
            for i in years:
                if str(i) not in (dict_.keys()):
                    dict_[str(i)]=0
            keys_=list(dict_.keys())
            for i in keys_:
                if(int(i) not in years):
                    del dict_[i]
            return {"scopus":dict_}
        except:
            
        # print(e.get_attribute("aria-label").split(" "))
            click_on_python=driver.find_elements(By.XPATH,f"""//*[local-name()='path'and @class="highcharts-halo highcharts-color-undefined" or local-name()='path'and@class="highcharts-point"]""")
            
            length1=0
            dict_={}
            for i in click_on_python:
                
                list_=i.get_attribute("aria-label").split(" ")
                dict_[list_[0][:4]]=list_[1]
                length1+=1

            print(length1)


            # dict_={}
            # for i in range(length1-4,length1+1):
            #     click_on_python=driver.find_element(By.XPATH,f"""//*[local-name()='path' and @class="highcharts-point"]{[i]}""")
            #     list_=click_on_python.get_attribute("aria-label").split(" ")
            #     dict_[list_[0][:4]]=list_[1]
            # print(dict_)
            for i in years:
                if str(i) not in (dict_.keys()):
                    dict_[str(i)]=0
            keys_=list(dict_.keys())
            for i in keys_:
                if(int(i) not in years):
                    del dict_[i]
            return {"scopus":dict_}
    
# print(citi_func_calling("https://www.scopus.com/authid/detail.uri?authorId=16635342600","https://www.webofscience.com/wos/author/rid/GNW-3254-2022","https://scholar.google.co.in/citations?user=575NT6IAAAAJ"))
def web_of_science_citi(url):
    if(url==None):
        for i in range(len(years)):
            empt_dict[str(years[i])]="0"
        return {"gs":empt_dict}
    else:
        url=f"https://www.webofscience.com/wos/author/rid/{url}"
        driver.get(url)
        e = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="onetrust-policy"]'))
            )

        click_ele=driver.find_element(By.XPATH,f"""//*[@id="onetrust-close-btn-container"]/button""")
        click_ele.click()
        e = WebDriverWait(driver, 40).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="dismissGuides"]'))
            )

        click_ele=driver.find_element(By.XPATH,f"""//*[@id="dismissGuides"]""")
        click_ele.click()
        dict_={}
        list_=[]
        for i in range(1,11):
            try:
                link=driver.find_element(By.XPATH,f"""//*[@id="mat-tab-content-0-0"]/div/div/div[2]/div/app-publication-card{[i]}/mat-card/app-article-metadata/div[1]/div[2]/div/div/a""")
                ActionChains(driver).move_to_element(link).key_down(Keys.COMMAND).click(link).key_up(Keys.COMMAND).perform()
                driver.switch_to.window(driver.window_handles[1])
                # print(driver.current_url)                  
                ele_=driver.find_elements(By.XPATH,"""//*[@id="filter-section-PY"]/div/div""")
                # print(ele_)
                for j in ele_:
                    list_.append(j.text)
                    # print(j.text)
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                time.sleep(3)   
            except:
                print(i)
                continue
        # print(list_)
        if(len(list_)==0):
            for i in range(len(years)):
                empt_dict[str(years[i])]="0"
            return {"wos":empt_dict} 
        else:

            s=""
            for i in list_:
                s=s+"\n"+i
            list_=s.split("\n")
            list_.remove("")
            for i in range(0,len(list_),2):
                if(list_[i] in dict_.keys()):
                    dict_[list_[i]]+=int(list_[i+1])
                else:
                    dict_[list_[i]]=int(list_[i+1])
            for i in years:
                if str(i) not in (dict_.keys()):
                    dict_[str(i)]=0
            keys_=list(dict_.keys())
            for i in keys_:
                if(int(i) not in years):
                    del dict_[i]
            return {"wos":dict_}




print(excel_read_write_fun("/Users/sundar.g/Downloads/Academic Identities of CSE Faculty.xlsx"))


    

  